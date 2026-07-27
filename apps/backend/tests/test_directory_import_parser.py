import io
import tempfile
import unittest
import zipfile
from pathlib import Path

from openpyxl import Workbook

from app.services.directory_import_parser import (
    DirectoryImportFormatError,
    DirectoryImportLimitError,
    ImportLimits,
    SourceRow,
    classify_cell_value,
    parse_directory_file,
    parse_legacy_layout,
    parse_table_sheet,
    safe_original_filename,
)


FIXTURES = Path(__file__).parent / "fixtures"


def limits(**overrides) -> ImportLimits:
    values = {
        "max_file_size_bytes": 10 * 1024 * 1024,
        "max_sheets": 20,
        "max_rows": 20000,
        "max_columns": 100,
        "max_cells": 200000,
        "max_cell_length": 2000,
        "max_zip_members": 1000,
        "max_uncompressed_bytes": 100 * 1024 * 1024,
    }
    values.update(overrides)
    return ImportLimits(**values)


class DirectoryImportParserTests(unittest.TestCase):
    def test_synthetic_legacy_fixture_produces_expected_candidates(self):
        parsed = parse_directory_file(
            FIXTURES / "directory_import_legacy.csv",
            original_filename="synthetic-directory.csv",
            parser_mode="legacy_layout",
            selected_sheet=None,
            column_mapping=None,
            limits=limits(),
        )

        people = [row for row in parsed.candidates if row.detected_kind == "person"]
        roles = [row for row in parsed.candidates if row.detected_kind == "role"]
        metadata = [
            row for row in parsed.candidates if row.detected_kind == "organization_metadata"
        ]

        self.assertEqual(len(people), 4)
        self.assertEqual(len(roles), 1)
        self.assertTrue(all(row.is_selected for row in people))
        self.assertFalse(any(row.is_selected for row in metadata))
        self.assertEqual(roles[0].normalized_data["work_phone"], "3 44 44")
        self.assertTrue(
            any(
                warning["code"] == "phone_type_uncertain"
                for warning in roles[0].warnings
            )
        )
        final_person = next(
            row for row in people if row.normalized_data["display_name"] == "Галина Образцова"
        )
        self.assertEqual(
            final_person.normalized_data["position"],
            "Приемная – главный специалист отдела",
        )
        self.assertEqual(final_person.normalized_data["email"], "contact@example.invalid")
        self.assertTrue(
            any(warning["code"] == "multiline_position" for warning in final_person.warnings)
        )

    def test_standard_table_detects_headers_and_accepts_manual_mapping(self):
        rows = [
            SourceRow("CSV", 1, ["Employee", "Telephone", "Team"]),
            SourceRow("CSV", 2, ["Test User", "+7 (000) 111-22-33", "IT"]),
        ]
        candidates, mapping, columns = parse_table_sheet(rows)
        self.assertEqual(
            mapping,
            {"0": "display_name", "1": "work_phone", "2": "department"},
        )
        self.assertEqual(candidates[0].normalized_data["display_name"], "Test User")
        self.assertEqual(columns[0]["samples"], ["Test User"])
        self.assertEqual(
            candidates[0].raw_cells["header_row"]["cells"],
            ["Employee", "Telephone", "Team"],
        )

        candidates, mapping, _ = parse_table_sheet(
            rows,
            column_mapping={"0": "display_name", "1": "internal_phone", "2": "position"},
        )
        self.assertEqual(mapping["1"], "internal_phone")
        self.assertEqual(candidates[0].normalized_data["position"], "IT")

        reconstructed, _, _ = parse_table_sheet(
            rows[1:],
            column_mapping={
                "0": "display_name",
                "1": "internal_phone",
                "2": "position",
            },
            source_has_header=False,
        )
        self.assertEqual(reconstructed[0].normalized_data["display_name"], "Test User")

    def test_standard_table_maps_full_internal_phone_header(self):
        rows = [
            SourceRow("CSV", 1, ["ФИО", "Внутренний номер"]),
            SourceRow("CSV", 2, ["Тестовый Контакт", "101"]),
        ]

        candidates, mapping, _ = parse_table_sheet(rows)

        self.assertEqual(mapping["1"], "internal_phone")
        self.assertEqual(candidates[0].normalized_data["internal_phone"], "101")

    def test_csv_supports_utf8_bom_semicolon_and_windows_1251(self):
        utf8 = "\ufeffФИО;Телефон\nТестовый Пользователь;3 10 10\n".encode("utf-8")
        cp1251 = "ФИО;Телефон\nДругой Пользователь;3 20 20\n".encode("cp1251")
        for payload, expected in (
            (utf8, "Тестовый Пользователь"),
            (cp1251, "Другой Пользователь"),
        ):
            with self.subTest(expected=expected), tempfile.NamedTemporaryFile(
                suffix=".csv"
            ) as source:
                source.write(payload)
                source.flush()
                parsed = parse_directory_file(
                    source.name,
                    original_filename="directory.csv",
                    parser_mode="table",
                    selected_sheet=None,
                    column_mapping=None,
                    limits=limits(),
                )
                self.assertEqual(parsed.candidates[0].normalized_data["display_name"], expected)

    def test_xlsx_formula_is_never_executed_and_is_reported(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Contacts"
        sheet.append(["Name", "Phone"])
        sheet.append(["Formula User", "=1+1"])
        with tempfile.NamedTemporaryFile(suffix=".xlsx") as source:
            workbook.save(source.name)
            workbook.close()
            parsed = parse_directory_file(
                source.name,
                original_filename="directory.xlsx",
                parser_mode="table",
                selected_sheet=None,
                column_mapping=None,
                limits=limits(),
            )
        candidate = parsed.candidates[0]
        self.assertIsNone(candidate.normalized_data["work_phone"])
        self.assertTrue(
            any(
                warning["code"] == "formula_without_cached_value"
                for warning in candidate.warnings
            )
        )
        self.assertFalse(
            any(
                warning["code"] == "formula_replaced_with_saved_value"
                for warning in candidate.warnings
            )
        )

    def test_rejects_spoofed_malformed_and_oversized_files(self):
        with tempfile.NamedTemporaryFile(suffix=".xlsx") as source:
            source.write(b"not a zip")
            source.flush()
            with self.assertRaises(DirectoryImportFormatError):
                parse_directory_file(
                    source.name,
                    original_filename="directory.xlsx",
                    parser_mode="auto",
                    selected_sheet=None,
                    column_mapping=None,
                    limits=limits(),
                )

        with tempfile.NamedTemporaryFile(suffix=".csv") as source:
            source.write(b"PK\x03\x04spoof")
            source.flush()
            with self.assertRaises(DirectoryImportFormatError):
                parse_directory_file(
                    source.name,
                    original_filename="directory.csv",
                    parser_mode="auto",
                    selected_sheet=None,
                    column_mapping=None,
                    limits=limits(),
                )

        with tempfile.NamedTemporaryFile(suffix=".csv") as source:
            source.write(b"Name,Phone\nTest User,12345\n")
            source.flush()
            with self.assertRaises(DirectoryImportLimitError):
                parse_directory_file(
                    source.name,
                    original_filename="directory.csv",
                    parser_mode="auto",
                    selected_sheet=None,
                    column_mapping=None,
                    limits=limits(max_file_size_bytes=5),
                )

    def test_value_classification_is_conservative(self):
        self.assertEqual(classify_cell_value("person@example.invalid"), "email")
        self.assertEqual(classify_cell_value("https://example.invalid"), "url")
        self.assertEqual(classify_cell_value("www.example.invalid"), "url")
        self.assertEqual(
            classify_cell_value("Тестовая область, район, пос. Примерный, д. 2"),
            "organization_metadata",
        )
        self.assertEqual(classify_cell_value("тел/факс 3 44 44"), "phone_fax")
        self.assertEqual(classify_cell_value("+7 (000) 111-22-33"), "phone")
        self.assertEqual(classify_cell_value("2026"), "text")
        self.assertEqual(classify_cell_value("Директор МО"), "position")
        self.assertEqual(classify_cell_value("Совет домов"), "text")
        self.assertEqual(
            safe_original_filename(r"..\private\directory.csv"),
            "directory.csv",
        )

    def test_table_classifies_email_in_phone_column_and_reports_truncation(self):
        rows = [
            SourceRow("CSV", 1, ["Name", "Phone"]),
            SourceRow("CSV", 2, ["Email User", "person@example.invalid"]),
            SourceRow(
                "CSV",
                3,
                ["Mixed User", "3 44 44; mixed@example.invalid"],
                truncated_columns=[0],
            ),
        ]
        candidates, _, _ = parse_table_sheet(rows)

        self.assertEqual(candidates[0].normalized_data["email"], "person@example.invalid")
        self.assertIsNone(candidates[0].normalized_data["work_phone"])
        self.assertTrue(
            any(item["code"] == "email_in_phone_column" for item in candidates[0].warnings)
        )
        self.assertEqual(candidates[1].normalized_data["work_phone"], "3 44 44")
        self.assertEqual(candidates[1].normalized_data["email"], "mixed@example.invalid")
        warning_codes = {item["code"] for item in candidates[1].warnings}
        self.assertIn("phone_and_email_combined", warning_codes)
        self.assertIn("value_truncated", warning_codes)

    def test_legacy_department_email_becomes_shared_contact(self):
        rows = [
            SourceRow("CSV", 1, ["Кластер Север", "", "shared@example.invalid"]),
            SourceRow("CSV", 2, ["Тестовый Сотрудник", "", "3 55 55"]),
        ]
        candidates = parse_legacy_layout(rows)
        shared = candidates[0]

        self.assertEqual(shared.detected_kind, "department_contact")
        self.assertEqual(shared.normalized_data["department"], "Кластер Север")
        self.assertEqual(shared.normalized_data["email"], "shared@example.invalid")
        self.assertTrue(
            any(item["code"] == "shared_department_contact" for item in shared.warnings)
        )
        self.assertEqual(candidates[1].normalized_data["department"], "Кластер Север")

    def test_csv_enforces_row_and_cell_limits(self):
        payload = b"Name,Phone\nOne User,12345\nTwo User,23456\n"
        with tempfile.NamedTemporaryFile(suffix=".csv") as source:
            source.write(payload)
            source.flush()
            with self.assertRaises(DirectoryImportLimitError):
                parse_directory_file(
                    source.name,
                    original_filename="directory.csv",
                    parser_mode="table",
                    selected_sheet=None,
                    column_mapping=None,
                    limits=limits(max_rows=2),
                )
            with self.assertRaises(DirectoryImportLimitError):
                parse_directory_file(
                    source.name,
                    original_filename="directory.csv",
                    parser_mode="table",
                    selected_sheet=None,
                    column_mapping=None,
                    limits=limits(max_cells=4),
                )

    def test_legacy_metadata_is_not_joined_to_first_person_and_phone_fax_is_preserved(self):
        rows = [
            SourceRow("Sheet", 1, ["Тестовая область, муниципальный район"]),
            SourceRow("Sheet", 2, ["пос. Примерный, д. 2"]),
            SourceRow("Sheet", 3, ["www.example.invalid"]),
            SourceRow("Sheet", 4, ["example.invalid"]),
            SourceRow("Sheet", 6, ["Директор"]),
            SourceRow("Sheet", 7, ["Анна Тестова", "6 12 500"]),
            SourceRow("Sheet", 9, ["Приемная"]),
            SourceRow(
                "Sheet",
                10,
                ["Тестовый Сотрудник", "тел/факс 3 44 44", "3 55 55"],
            ),
        ]

        candidates = parse_legacy_layout(rows)
        metadata = [
            candidate
            for candidate in candidates
            if candidate.detected_kind == "organization_metadata"
        ]
        people = [
            candidate for candidate in candidates if candidate.detected_kind == "person"
        ]

        self.assertEqual(
            [(candidate.source_row_start, candidate.source_row_end) for candidate in metadata],
            [(1, 1), (2, 2), (3, 3), (4, 4)],
        )
        self.assertTrue(all(not candidate.is_selected for candidate in metadata))
        self.assertEqual((people[0].source_row_start, people[0].source_row_end), (6, 7))
        self.assertEqual(people[0].normalized_data["work_phone"], "6 12 500")
        self.assertIn(
            "phone_type_uncertain",
            {warning["code"] for warning in people[0].warnings},
        )
        self.assertEqual(people[1].normalized_data["work_phone"], "тел/факс 3 44 44")
        self.assertIsNone(people[1].normalized_data["internal_phone"])
        self.assertIn(
            "phone_fax_source",
            {warning["code"] for warning in people[1].warnings},
        )
        self.assertIn(
            "multiple_phone_values",
            {warning["code"] for warning in people[1].warnings},
        )

    def test_csv_handles_quoted_values_and_rejects_malformed_or_binary_input(self):
        payload = (
            'Name,Notes\r\n"Test User","line one, quoted\r\nline two"\r\n'
            '"Formula Text","=1+1"\r\n'
        ).encode()
        with tempfile.NamedTemporaryFile(suffix=".csv") as source:
            source.write(payload)
            source.flush()
            parsed = parse_directory_file(
                source.name,
                original_filename="directory.csv",
                parser_mode="table",
                selected_sheet=None,
                column_mapping=None,
                limits=limits(),
            )
        self.assertEqual(
            parsed.candidates[0].normalized_data["notes"].splitlines(),
            ["line one, quoted", "line two"],
        )
        self.assertEqual(parsed.candidates[1].normalized_data["notes"], "=1+1")

        for payload in (b'Name,Notes\n"unterminated,value\n', b"Name,Phone\nTest,\x01\x02\x03"):
            with self.subTest(payload=payload), tempfile.NamedTemporaryFile(
                suffix=".csv"
            ) as source:
                source.write(payload)
                source.flush()
                with self.assertRaises(DirectoryImportFormatError):
                    parse_directory_file(
                        source.name,
                        original_filename="directory.csv",
                        parser_mode="table",
                        selected_sheet=None,
                        column_mapping=None,
                        limits=limits(),
                    )

    def test_xlsx_macro_content_type_is_rejected(self):
        workbook = Workbook()
        workbook.active.append(["Name", "Phone"])
        workbook.active.append(["Test User", "3 11 11"])
        with (
            tempfile.NamedTemporaryFile(suffix=".xlsx") as source,
            tempfile.NamedTemporaryFile(suffix=".xlsx") as modified,
        ):
            workbook.save(source.name)
            workbook.close()
            with (
                zipfile.ZipFile(source.name) as source_archive,
                zipfile.ZipFile(modified.name, "w") as target_archive,
            ):
                for member in source_archive.infolist():
                    value = source_archive.read(member)
                    if member.filename == "[Content_Types].xml":
                        value = value.replace(
                            b"application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml",
                            b"application/vnd.ms-excel.sheet.macroEnabled.main+xml",
                        )
                    target_archive.writestr(member, value)
            with self.assertRaises(DirectoryImportFormatError):
                parse_directory_file(
                    modified.name,
                    original_filename="directory.xlsx",
                    parser_mode="table",
                    selected_sheet=None,
                    column_mapping=None,
                    limits=limits(),
                )

    def test_only_replayable_worksheets_are_exposed(self):
        workbook = Workbook()
        empty = workbook.active
        empty.title = "Empty"
        contacts = workbook.create_sheet("Contacts")
        contacts.append(["Name", "Phone"])
        contacts.append(["Test User", "3 11 11"])
        with tempfile.NamedTemporaryFile(suffix=".xlsx") as source:
            workbook.save(source.name)
            workbook.close()
            parsed = parse_directory_file(
                source.name,
                original_filename="directory.xlsx",
                parser_mode="auto",
                selected_sheet=None,
                column_mapping=None,
                limits=limits(),
            )

        self.assertEqual(parsed.available_sheets, ["Contacts"])
        self.assertEqual(parsed.selected_sheet, "Contacts")


if __name__ == "__main__":
    unittest.main()
