import csv
import hashlib
import io
import re
import zipfile
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Literal

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

ParserMode = Literal["auto", "table", "legacy_layout"]
DetectedKind = Literal[
    "person", "role", "department_contact", "organization_metadata", "unknown"
]

TARGET_FIELDS = (
    "display_name",
    "department",
    "position",
    "internal_phone",
    "work_phone",
    "mobile_phone",
    "email",
    "room",
    "location",
    "notes",
)

HEADER_ALIASES = {
    "display_name": {"фио", "имя", "сотрудник", "name", "full name", "employee"},
    "department": {"отдел", "подразделение", "department", "unit", "team"},
    "position": {"должность", "position", "job title", "title"},
    "internal_phone": {
        "внутренний",
        "внутренний номер",
        "внутренний телефон",
        "добавочный",
        "extension",
        "ext",
        "internal phone",
    },
    "work_phone": {"телефон", "рабочий телефон", "phone", "telephone", "work phone"},
    "mobile_phone": {"мобильный", "мобильный телефон", "mobile", "cell"},
    "email": {"email", "e-mail", "почта"},
    "room": {"кабинет", "комната", "room", "office"},
    "location": {"расположение", "адрес", "location"},
    "notes": {"примечания", "комментарий", "notes", "comment"},
}
ROLE_WORDS = {
    "директор",
    "секретарь",
    "глава",
    "заместитель",
    "специалист",
    "руководитель",
    "начальник",
    "главный",
    "приемная",
    "director",
    "secretary",
    "head",
    "deputy",
    "specialist",
    "manager",
}
DEPARTMENT_WORDS = {"отдел", "кластер", "управление", "department", "cluster", "division"}
METADATA_WORDS = {"адрес", "сайт", "организация", "organization", "website", "address"}
EMAIL_RE = re.compile(r"^[^@\s]{1,64}@[^@\s.]{1,190}\.[^@\s]{2,63}$", re.IGNORECASE)
EMAIL_SEARCH_RE = re.compile(r"[^@\s,;]{1,64}@[^@\s,;.]{1,190}\.[^@\s,;]{2,63}", re.IGNORECASE)
URL_RE = re.compile(r"^https?://[^\s]+$", re.IGNORECASE)
DOMAIN_RE = re.compile(
    r"^(?:www\.)?[a-z0-9](?:[a-z0-9-]{0,61}[a-z0-9])?"
    r"(?:\.[a-z0-9](?:[a-z0-9-]{0,61}[a-z0-9])?)+(?:/[^\s]*)?$",
    re.IGNORECASE,
)
ADDRESS_RE = re.compile(
    r"\b(?:область|район|пос(?:елок)?\.?|квартал|улица|ул\.|дом|д\.)\b",
    re.IGNORECASE,
)
PHONE_PREFIX_RE = re.compile(r"^\s*(?:тел(?:ефон)?|факс|тел\s*/\s*факс|phone|fax)\b", re.IGNORECASE)
CONTROL_RE = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f]")


class DirectoryImportError(ValueError):
    code = "invalid_file"


class DirectoryImportLimitError(DirectoryImportError):
    code = "limit_exceeded"


class DirectoryImportFormatError(DirectoryImportError):
    code = "unsupported_or_spoofed_format"


@dataclass(frozen=True)
class ImportLimits:
    max_file_size_bytes: int
    max_sheets: int
    max_rows: int
    max_columns: int
    max_cells: int
    max_cell_length: int
    max_zip_members: int
    max_uncompressed_bytes: int


@dataclass
class SourceRow:
    sheet: str | None
    row_number: int
    cells: list[str]
    formula_columns: list[int] = field(default_factory=list)
    formula_without_cached_value_columns: list[int] = field(default_factory=list)
    truncated_columns: list[int] = field(default_factory=list)


@dataclass
class ImportCandidate:
    source_sheet: str | None
    source_row_start: int
    source_row_end: int
    raw_cells: dict[str, Any]
    detected_kind: DetectedKind
    confidence: float | None
    normalized_data: dict[str, str | None]
    warnings: list[dict[str, Any]]
    is_selected: bool
    proposed_action: Literal["create", "skip"]
    sort_order: int = 0


@dataclass
class ParsedDirectoryFile:
    file_type: Literal["xlsx", "csv"]
    file_sha256: str
    available_sheets: list[str]
    selected_sheet: str | None
    parser_mode: ParserMode
    detected_mode: Literal["table", "legacy_layout"]
    column_mapping: dict[str, str]
    source_columns: list[dict[str, Any]]
    total_source_rows: int
    candidates: list[ImportCandidate]
    global_warnings: list[dict[str, Any]]


def safe_original_filename(value: str | None) -> str:
    name = (value or "directory-import").replace("\\", "/").rsplit("/", 1)[-1]
    name = CONTROL_RE.sub("", name).strip()
    return (name or "directory-import")[:255]


def classify_cell_value(value: str) -> str:
    normalized = value.strip()
    if not normalized:
        return "empty"
    if EMAIL_RE.fullmatch(normalized):
        return "email"
    if URL_RE.fullmatch(normalized) or DOMAIN_RE.fullmatch(normalized):
        return "url"
    digits = re.sub(r"\D", "", normalized)
    prefixed = bool(PHONE_PREFIX_RE.match(normalized))
    if EMAIL_SEARCH_RE.search(normalized) and 3 <= len(digits) <= 20:
        return "phone_email"
    if 3 <= len(digits) <= 20 and (prefixed or len(digits) >= 5):
        return "phone_fax" if "факс" in normalized.lower() or "fax" in normalized.lower() else "phone"
    if (
        URL_RE.search(normalized)
        or DOMAIN_RE.fullmatch(normalized)
        or _contains_word(normalized, METADATA_WORDS)
        or ADDRESS_RE.search(normalized)
    ):
        return "organization_metadata"
    if _contains_word(normalized, DEPARTMENT_WORDS):
        return "department"
    if _contains_word(normalized, ROLE_WORDS):
        return "position"
    if _looks_like_person(normalized):
        return "person"
    return "text"


def normalize_candidate(data: dict[str, Any]) -> dict[str, str | None]:
    result: dict[str, str | None] = {}
    for field_name in TARGET_FIELDS:
        raw = data.get(field_name)
        value = str(raw).strip() if raw is not None else ""
        result[field_name] = value or None
    if result["email"]:
        result["email"] = result["email"].lower()
    return result


def detect_parser_mode(rows: list[SourceRow]) -> Literal["table", "legacy_layout"]:
    _, mapping, _ = detect_table_header(rows)
    return "table" if len(mapping) >= 2 and "display_name" in mapping.values() else "legacy_layout"


def detect_table_header(
    rows: list[SourceRow],
) -> tuple[int | None, dict[str, str], list[dict[str, Any]]]:
    best_row: SourceRow | None = None
    best_mapping: dict[str, str] = {}
    for row in rows[:30]:
        mapping: dict[str, str] = {}
        assigned: set[str] = set()
        for index, value in enumerate(row.cells):
            normalized = _normalize_header(value)
            matches = [
                field_name
                for field_name, aliases in HEADER_ALIASES.items()
                if normalized in aliases and field_name not in assigned
            ]
            if matches:
                mapping[str(index)] = matches[0]
                assigned.add(matches[0])
        if len(mapping) > len(best_mapping):
            best_row = row
            best_mapping = mapping
    if best_row is None:
        return None, {}, []
    source_columns = [
        {
            "index": index,
            "label": value,
            "samples": _column_samples(rows, index, best_row.row_number),
        }
        for index, value in enumerate(best_row.cells)
    ]
    return best_row.row_number, best_mapping, source_columns


def parse_table_sheet(
    rows: list[SourceRow],
    *,
    column_mapping: dict[str, str] | None = None,
    source_has_header: bool = True,
) -> tuple[list[ImportCandidate], dict[str, str], list[dict[str, Any]]]:
    detected_header_row, automatic_mapping, source_columns = detect_table_header(rows)
    if column_mapping:
        header_row = detected_header_row if source_has_header else None
        mapping = _validated_mapping(column_mapping)
    else:
        header_row = detected_header_row
        mapping = _validated_mapping(automatic_mapping)
    candidates: list[ImportCandidate] = []
    for row in rows:
        if header_row is not None and row.row_number <= header_row:
            continue
        if not any(row.cells):
            continue
        normalized: dict[str, Any] = {}
        warnings = _formula_warnings(row)
        for source_index, target_field in mapping.items():
            index = int(source_index)
            value = row.cells[index] if index < len(row.cells) else ""
            if target_field in {"internal_phone", "work_phone", "mobile_phone"}:
                email_match = EMAIL_SEARCH_RE.search(value)
                if email_match:
                    normalized["email"] = email_match.group(0)
                    remaining = value.replace(email_match.group(0), " ").strip(" ,;/")
                    if remaining:
                        normalized[target_field] = remaining
                        if not any(
                            item["code"] == "phone_and_email_combined"
                            for item in warnings
                        ):
                            warnings.append(
                                _warning("phone_and_email_combined", "warning")
                            )
                    else:
                        warnings.append(_warning("email_in_phone_column", "warning"))
                    continue
            normalized[target_field] = value
        normalized_data = normalize_candidate(normalized)
        warnings.extend(
            _phone_warnings(
                value
                for field_name, value in normalized_data.items()
                if field_name in {"internal_phone", "work_phone", "mobile_phone"} and value
            )
        )
        has_name = bool(normalized_data["display_name"])
        if not has_name:
            warnings.append(_warning("missing_display_name", "blocking"))
        if not mapping:
            warnings.append(_warning("column_mapping_required", "blocking"))
        candidates.append(
            ImportCandidate(
                source_sheet=row.sheet,
                source_row_start=row.row_number,
                source_row_end=row.row_number,
                raw_cells=_raw_rows([row]),
                detected_kind="person" if has_name else "unknown",
                confidence=0.95 if has_name and mapping else 0.25,
                normalized_data=normalized_data,
                warnings=warnings,
                is_selected=has_name and bool(mapping),
                proposed_action="create" if has_name and mapping else "skip",
            )
        )
    if detected_header_row is not None and candidates:
        header_source = next(
            (row for row in rows if row.row_number == detected_header_row),
            None,
        )
        if header_source is not None:
            candidates[0].raw_cells["header_row"] = _raw_row(header_source)
    return candidates, mapping, source_columns


def parse_legacy_layout(rows: list[SourceRow]) -> list[ImportCandidate]:
    candidates: list[ImportCandidate] = []
    pending_position: list[SourceRow] = []
    pending_contacts: list[str] = []
    current_department: str | None = None

    def flush_pending_unknown() -> None:
        nonlocal pending_position, pending_contacts
        if not pending_position:
            return
        text = " ".join(_non_contact_text(row.cells) for row in pending_position).strip()
        warnings = [_warning("unknown_row_structure", "warning")]
        if len(pending_position) > 1:
            warnings.append(_warning("multiline_position", "info"))
        candidates.append(
            _candidate(
                pending_position,
                "unknown",
                {"display_name": text, "position": text, "department": current_department},
                warnings,
                selected=False,
                confidence=0.3,
            )
        )
        pending_position = []
        pending_contacts = []

    for row in rows:
        values = [value for value in row.cells if value]
        if not values:
            continue
        kinds = [classify_cell_value(value) for value in values]
        person_index = next((index for index, kind in enumerate(kinds) if kind == "person"), None)
        contacts = [
            value
            for value, kind in zip(values, kinds, strict=True)
            if kind in {"email", "phone", "phone_fax", "phone_email"}
        ]
        text_values = [
            value
            for value, kind in zip(values, kinds, strict=True)
            if kind not in {"email", "phone", "phone_fax", "phone_email", "empty"}
        ]

        if person_index is not None:
            person = values[person_index]
            position = " ".join(
                text for text in (_non_contact_text(item.cells) for item in pending_position) if text
            ).strip()
            warnings = _formula_warnings(row)
            if len(pending_position) > 1:
                warnings.append(_warning("multiline_position", "info"))
            all_contacts = [*pending_contacts, *contacts]
            data = _contacts_to_data(all_contacts)
            data.update(
                {
                    "display_name": person,
                    "position": position or None,
                    "department": current_department,
                }
            )
            candidates.append(
                _candidate(
                    [*pending_position, row],
                    "person",
                    data,
                    warnings,
                    selected=True,
                    confidence=0.88,
                )
            )
            pending_position = []
            pending_contacts = []
            continue

        department_text = next(
            (value for value, kind in zip(values, kinds, strict=True) if kind == "department"),
            None,
        )
        if department_text:
            flush_pending_unknown()
            current_department = department_text
            if contacts:
                data = _contacts_to_data(contacts)
                data.update(
                    {
                        "display_name": department_text,
                        "department": department_text,
                    }
                )
                candidates.append(
                    _candidate(
                        [row],
                        "department_contact",
                        data,
                        [
                            *_formula_warnings(row),
                            _warning("shared_department_contact", "warning"),
                        ],
                        selected=True,
                        confidence=0.8,
                    )
                )
            else:
                candidates.append(
                    _candidate(
                        [row],
                        "organization_metadata",
                        {"display_name": department_text, "department": department_text},
                        [_warning("department_context", "info")],
                        selected=False,
                        confidence=0.75,
                    )
                )
            continue

        if any(kind in {"url", "organization_metadata"} for kind in kinds):
            flush_pending_unknown()
            candidates.append(
                _candidate(
                    [row],
                    "organization_metadata",
                    {"display_name": " ".join(values), "notes": " ".join(values)},
                    [_warning("organization_metadata", "info"), *_formula_warnings(row)],
                    selected=False,
                    confidence=0.9,
                )
            )
            continue

        if contacts and not text_values and pending_position:
            data = _contacts_to_data([*pending_contacts, *contacts])
            position = " ".join(_non_contact_text(item.cells) for item in pending_position).strip()
            data.update(
                {
                    "display_name": position,
                    "position": position,
                    "department": current_department,
                }
            )
            warnings = [_warning("role_without_person", "warning")]
            if len(pending_position) > 1:
                warnings.append(_warning("multiline_position", "info"))
            candidates.append(
                _candidate(
                    [*pending_position, row],
                    "role",
                    data,
                    warnings,
                    selected=True,
                    confidence=0.72,
                )
            )
            pending_position = []
            pending_contacts = []
            continue

        if text_values:
            pending_position.append(row)
            pending_contacts.extend(contacts)
            continue

        candidates.append(
            _candidate(
                [row],
                "unknown",
                _contacts_to_data(contacts),
                [_warning("missing_display_name", "blocking"), *_formula_warnings(row)],
                selected=False,
                confidence=0.2,
            )
        )

    flush_pending_unknown()
    return candidates


def parse_directory_file(
    path: str | Path,
    *,
    original_filename: str,
    parser_mode: ParserMode,
    selected_sheet: str | None,
    column_mapping: dict[str, str] | None,
    limits: ImportLimits,
) -> ParsedDirectoryFile:
    file_path = Path(path)
    extension = Path(original_filename).suffix.lower()
    if extension not in {".xlsx", ".csv"}:
        raise DirectoryImportFormatError("Only XLSX and CSV files are supported")
    if file_path.stat().st_size > limits.max_file_size_bytes:
        raise DirectoryImportLimitError("File is too large")
    file_hash = _sha256_file(file_path)
    global_warnings: list[dict[str, Any]] = []
    if extension == ".xlsx":
        sheets, global_warnings = _read_xlsx(file_path, limits)
        file_type: Literal["xlsx", "csv"] = "xlsx"
    else:
        sheets = {"CSV": _read_csv(file_path, limits)}
        file_type = "csv"

    available_sheets = list(sheets)
    if not available_sheets:
        raise DirectoryImportFormatError("The file does not contain readable sheets")
    if selected_sheet is not None and selected_sheet not in sheets:
        raise DirectoryImportError("Selected worksheet was not found")
    chosen_sheet = selected_sheet or available_sheets[0]
    chosen_rows = sheets[chosen_sheet]
    detected_mode = detect_parser_mode(chosen_rows)
    effective_mode = detected_mode if parser_mode == "auto" else parser_mode
    all_candidates: list[ImportCandidate] = []
    selected_mapping: dict[str, str] = {}
    source_columns: list[dict[str, Any]] = []
    for sheet_name, rows in sheets.items():
        mode = detect_parser_mode(rows) if parser_mode == "auto" else parser_mode
        if mode == "table":
            candidates, mapping, columns = parse_table_sheet(
                rows, column_mapping=column_mapping if sheet_name == chosen_sheet else None
            )
            if sheet_name == chosen_sheet:
                selected_mapping = mapping
                source_columns = columns
        else:
            candidates = parse_legacy_layout(rows)
            if sheet_name == chosen_sheet:
                _, selected_mapping, source_columns = detect_table_header(rows)
        if sheet_name != chosen_sheet:
            for candidate in candidates:
                candidate.is_selected = False
                candidate.proposed_action = "skip"
        all_candidates.extend(candidates)
    for index, candidate in enumerate(all_candidates):
        candidate.sort_order = index
    replayable_sheets = [
        sheet_name
        for sheet_name in available_sheets
        if any(candidate.source_sheet == sheet_name for candidate in all_candidates)
    ]
    if not replayable_sheets:
        raise DirectoryImportError("The file does not contain import preview rows")
    if chosen_sheet not in replayable_sheets:
        if selected_sheet is not None:
            raise DirectoryImportError("Selected worksheet has no import preview rows")
        return parse_directory_file(
            file_path,
            original_filename=original_filename,
            parser_mode=parser_mode,
            selected_sheet=replayable_sheets[0],
            column_mapping=column_mapping,
            limits=limits,
        )
    if global_warnings:
        first_selected_candidate = next(
            (
                candidate
                for candidate in all_candidates
                if candidate.source_sheet == chosen_sheet
            ),
            None,
        )
        if first_selected_candidate is not None:
            for warning in global_warnings:
                if warning not in first_selected_candidate.warnings:
                    first_selected_candidate.warnings.append(warning)
    return ParsedDirectoryFile(
        file_type=file_type,
        file_sha256=file_hash,
        available_sheets=replayable_sheets,
        selected_sheet=chosen_sheet,
        parser_mode=parser_mode,
        detected_mode=effective_mode,
        column_mapping=selected_mapping,
        source_columns=source_columns,
        total_source_rows=sum(len(rows) for rows in sheets.values()),
        candidates=all_candidates,
        global_warnings=global_warnings,
    )


def _read_xlsx(
    path: Path, limits: ImportLimits
) -> tuple[dict[str, list[SourceRow]], list[dict[str, Any]]]:
    warnings = _inspect_xlsx_archive(path, limits)
    formula_book = None
    value_book = None
    try:
        formula_book = load_workbook(
            path, read_only=True, data_only=False, keep_links=False, rich_text=False
        )
        value_book = load_workbook(
            path, read_only=True, data_only=True, keep_links=False, rich_text=False
        )
    except (OSError, ValueError, KeyError, zipfile.BadZipFile, InvalidFileException) as exc:
        if formula_book is not None:
            formula_book.close()
        if value_book is not None:
            value_book.close()
        raise DirectoryImportFormatError("Malformed or protected XLSX file") from exc
    try:
        if len(formula_book.sheetnames) > limits.max_sheets:
            raise DirectoryImportLimitError("Workbook has too many sheets")
        result: dict[str, list[SourceRow]] = {}
        total_rows = 0
        total_cells = 0
        for sheet_name in formula_book.sheetnames:
            formula_sheet = formula_book[sheet_name]
            value_sheet = value_book[sheet_name]
            sheet_rows: list[SourceRow] = []
            value_iterator = value_sheet.iter_rows()
            for row_number, formula_row in enumerate(formula_sheet.iter_rows(), start=1):
                value_row = next(value_iterator, ())
                total_rows += 1
                if total_rows > limits.max_rows:
                    raise DirectoryImportLimitError("Workbook has too many rows")
                if len(formula_row) > limits.max_columns:
                    raise DirectoryImportLimitError("Workbook has too many columns")
                total_cells += len(formula_row)
                if total_cells > limits.max_cells:
                    raise DirectoryImportLimitError("Workbook has too many cells")
                cells: list[str] = []
                formula_columns: list[int] = []
                formula_without_cached_value_columns: list[int] = []
                truncated_columns: list[int] = []
                for index, cell in enumerate(formula_row):
                    value = cell.value
                    if cell.data_type == "f" or (isinstance(value, str) and value.startswith("=")):
                        formula_columns.append(index)
                        value = value_row[index].value if index < len(value_row) else None
                        if value is None:
                            formula_without_cached_value_columns.append(index)
                    safe_value, was_truncated = _safe_cell(value, limits.max_cell_length)
                    cells.append(safe_value)
                    if was_truncated:
                        truncated_columns.append(index)
                sheet_rows.append(
                    SourceRow(
                        sheet=sheet_name,
                        row_number=row_number,
                        cells=_trim_trailing_empty(cells),
                        formula_columns=formula_columns,
                        formula_without_cached_value_columns=formula_without_cached_value_columns,
                        truncated_columns=truncated_columns,
                    )
                )
            result[sheet_name] = sheet_rows
        return result, warnings
    finally:
        formula_book.close()
        value_book.close()


def _inspect_xlsx_archive(path: Path, limits: ImportLimits) -> list[dict[str, Any]]:
    try:
        with zipfile.ZipFile(path) as archive:
            members = archive.infolist()
            if len(members) > limits.max_zip_members:
                raise DirectoryImportLimitError("XLSX archive has too many members")
            total_size = sum(member.file_size for member in members)
            if total_size > limits.max_uncompressed_bytes:
                raise DirectoryImportLimitError("XLSX archive expands beyond the configured limit")
            names = {member.filename for member in members}
            if "[Content_Types].xml" not in names or "xl/workbook.xml" not in names:
                raise DirectoryImportFormatError("The uploaded file is not an XLSX workbook")
            if any(member.flag_bits & 0x1 for member in members):
                raise DirectoryImportFormatError("Encrypted XLSX files are not supported")
            if any(name.lower().endswith("vbaproject.bin") for name in names):
                raise DirectoryImportFormatError("Macro-enabled workbooks are not supported")
            if archive.getinfo("[Content_Types].xml").file_size > 1024 * 1024:
                raise DirectoryImportFormatError("Malformed XLSX content types")
            content_types = archive.read("[Content_Types].xml")
            if b"macroEnabled" in content_types or b"vbaProject" in content_types:
                raise DirectoryImportFormatError("Macro-enabled workbooks are not supported")
            for member in members:
                if member.file_size and not member.compress_size:
                    raise DirectoryImportLimitError("Suspicious XLSX compression ratio")
                if (
                    member.file_size
                    and member.compress_size
                    and member.file_size / member.compress_size > 1000
                ):
                    raise DirectoryImportLimitError("Suspicious XLSX compression ratio")
            warnings = []
            if any(name.startswith("xl/externalLinks/") for name in names):
                warnings.append(_warning("external_links_ignored", "warning"))
            return warnings
    except zipfile.BadZipFile as exc:
        raise DirectoryImportFormatError("The uploaded file is not a valid XLSX ZIP container") from exc


def _read_csv(path: Path, limits: ImportLimits) -> list[SourceRow]:
    raw = path.read_bytes()
    if raw.startswith(b"PK\x03\x04"):
        raise DirectoryImportFormatError("CSV extension does not match the uploaded format")
    if any(byte < 0x20 and byte not in {0x09, 0x0A, 0x0D} for byte in raw):
        raise DirectoryImportFormatError("CSV contains binary data")
    try:
        text = raw.decode("utf-8-sig")
    except UnicodeDecodeError:
        try:
            text = raw.decode("cp1251")
        except UnicodeDecodeError as exc:
            raise DirectoryImportFormatError("CSV must use UTF-8 or Windows-1251") from exc
    sample = text[:8192]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t")
        delimiter = dialect.delimiter
    except csv.Error:
        delimiter = ";" if sample.count(";") >= sample.count(",") else ","
    previous_limit = csv.field_size_limit()
    csv.field_size_limit(limits.max_cell_length * 4)
    rows: list[SourceRow] = []
    total_cells = 0
    try:
        for row_number, values in enumerate(
            csv.reader(io.StringIO(text), delimiter=delimiter, strict=True),
            start=1,
        ):
            if row_number > limits.max_rows:
                raise DirectoryImportLimitError("CSV has too many rows")
            if len(values) > limits.max_columns:
                raise DirectoryImportLimitError("CSV has too many columns")
            total_cells += len(values)
            if total_cells > limits.max_cells:
                raise DirectoryImportLimitError("CSV has too many cells")
            cells: list[str] = []
            truncated_columns: list[int] = []
            for index, value in enumerate(values):
                safe_value, was_truncated = _safe_cell(value, limits.max_cell_length)
                cells.append(safe_value)
                if was_truncated:
                    truncated_columns.append(index)
            rows.append(
                SourceRow(
                    sheet="CSV",
                    row_number=row_number,
                    cells=_trim_trailing_empty(cells),
                    truncated_columns=truncated_columns,
                )
            )
    except csv.Error as exc:
        raise DirectoryImportFormatError("Malformed CSV file") from exc
    finally:
        csv.field_size_limit(previous_limit)
    return rows


def _candidate(
    rows: list[SourceRow],
    kind: DetectedKind,
    data: dict[str, Any],
    warnings: list[dict[str, Any]],
    *,
    selected: bool,
    confidence: float,
) -> ImportCandidate:
    combined_warnings = list(warnings)
    source_phone_values = [
        value
        for row in rows
        for value in row.cells
        if classify_cell_value(value) in {"phone", "phone_fax"}
    ]
    combined_warnings.extend(
        warning
        for warning in _phone_warnings(source_phone_values)
        if warning not in combined_warnings
    )
    if len(source_phone_values) > 1:
        combined_warnings.append(_warning("multiple_phone_values", "warning"))
    for row in rows:
        for warning in _formula_warnings(row):
            if warning not in combined_warnings:
                combined_warnings.append(warning)
    return ImportCandidate(
        source_sheet=rows[0].sheet,
        source_row_start=rows[0].row_number,
        source_row_end=rows[-1].row_number,
        raw_cells=_raw_rows(rows),
        detected_kind=kind,
        confidence=confidence,
        normalized_data=normalize_candidate(data),
        warnings=combined_warnings,
        is_selected=selected
        and not any(item["severity"] == "blocking" for item in combined_warnings),
        proposed_action="create" if selected else "skip",
    )


def _contacts_to_data(values: list[str]) -> dict[str, str | None]:
    result: dict[str, str | None] = {}
    phones: list[str] = []
    for value in values:
        kind = classify_cell_value(value)
        if kind == "email" and "email" not in result:
            result["email"] = value
        elif kind in {"phone", "phone_fax"}:
            phones.append(value)
        elif kind == "phone_email":
            email_match = EMAIL_SEARCH_RE.search(value)
            if email_match and "email" not in result:
                result["email"] = email_match.group(0)
            if email_match:
                phone_value = value.replace(email_match.group(0), " ").strip(" ,;/")
                if phone_value:
                    phones.append(phone_value)
    if phones:
        result["work_phone"] = phones[0]
    return result


def _raw_rows(rows: list[SourceRow]) -> dict[str, Any]:
    return {"rows": [_raw_row(row) for row in rows]}


def _raw_row(row: SourceRow) -> dict[str, Any]:
    return {
        "row": row.row_number,
        "cells": row.cells,
        "formula_columns": row.formula_columns,
        "formula_without_cached_value_columns": row.formula_without_cached_value_columns,
        "truncated_columns": row.truncated_columns,
    }


def _formula_warnings(row: SourceRow) -> list[dict[str, Any]]:
    warnings: list[dict[str, Any]] = []
    saved_value_columns = [
        column
        for column in row.formula_columns
        if column not in row.formula_without_cached_value_columns
    ]
    if saved_value_columns:
        warnings.append(
            _warning(
                "formula_replaced_with_saved_value",
                "warning",
                columns=saved_value_columns,
            )
        )
    if row.formula_without_cached_value_columns:
        warnings.append(
            _warning(
                "formula_without_cached_value",
                "warning",
                columns=row.formula_without_cached_value_columns,
            )
        )
    if row.truncated_columns:
        warnings.append(
            _warning("value_truncated", "warning", columns=row.truncated_columns)
        )
    for value in row.cells:
        if classify_cell_value(value) == "phone_email":
            warnings.append(_warning("phone_and_email_combined", "warning"))
            break
    return warnings


def _phone_warnings(values: Any) -> list[dict[str, Any]]:
    warnings: list[dict[str, Any]] = []
    for value in values:
        normalized = str(value).strip()
        digits = re.sub(r"\D", "", normalized)
        if not digits:
            continue
        if len(digits) < 10 and not normalized.startswith("+"):
            warning = _warning(
                "phone_type_uncertain",
                "warning",
                normalized_digits=digits,
            )
            if warning not in warnings:
                warnings.append(warning)
        if classify_cell_value(normalized) == "phone_fax":
            warning = _warning("phone_fax_source", "warning")
            if warning not in warnings:
                warnings.append(warning)
    return warnings


def _warning(code: str, severity: str, **details: Any) -> dict[str, Any]:
    return {"code": code, "severity": severity, **details}


def _validated_mapping(mapping: dict[str, str]) -> dict[str, str]:
    result: dict[str, str] = {}
    assigned: set[str] = set()
    for source, target in mapping.items():
        if not str(source).isdigit() or target not in TARGET_FIELDS or target in assigned:
            continue
        result[str(int(source))] = target
        assigned.add(target)
    return result


def _normalize_header(value: str) -> str:
    return re.sub(r"\s+", " ", value.strip().lower().replace("ё", "е"))


def _column_samples(rows: list[SourceRow], index: int, header_row: int) -> list[str]:
    samples: list[str] = []
    for row in rows:
        if row.row_number <= header_row or index >= len(row.cells) or not row.cells[index]:
            continue
        samples.append(row.cells[index])
        if len(samples) == 3:
            break
    return samples


def _looks_like_person(value: str) -> bool:
    if any(char.isdigit() for char in value) or _contains_word(value, ROLE_WORDS | DEPARTMENT_WORDS):
        return False
    words = re.findall(r"[A-Za-zА-Яа-яЁё-]+", value)
    return (
        2 <= len(words) <= 4
        and all(len(word.strip("-")) >= 2 for word in words)
        and all(word[0].isupper() for word in words)
        and len(" ".join(words)) >= 6
    )


def _contains_word(value: str, words: set[str]) -> bool:
    lowered = value.lower().replace("ё", "е")
    return any(re.search(rf"\b{re.escape(word)}\b", lowered) for word in words)


def _non_contact_text(values: list[str]) -> str:
    return " ".join(
        value
        for value in values
        if classify_cell_value(value)
        not in {"email", "phone", "phone_fax", "phone_email", "empty"}
    ).strip()


def _safe_cell(value: Any, max_length: int) -> tuple[str, bool]:
    if value is None:
        return "", False
    text = CONTROL_RE.sub("", str(value)).strip()
    return text[:max_length], len(text) > max_length


def _trim_trailing_empty(values: list[str]) -> list[str]:
    while values and not values[-1]:
        values.pop()
    return values


def _sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as source:
        for chunk in iter(lambda: source.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()
